import calendar
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins

ANO = 2026

# Função para gerar a lista de datas do ano
def generate_dates(year=2025):
    dias_semana = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira', 'Sábado', 'Domingo']
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    dates = []
    for month in range(1, 13):
        num_days = calendar.monthrange(year, month)[1]
        for day in range(1, num_days + 1):
            weekday = dias_semana[calendar.weekday(year, month, day)]
            month_name = meses[month - 1]
            date = f"{day} | {weekday} | {month_name} | {year}"
            dates.append(date)
    return dates

# Função para configurar a página de cada dia
def setup_day_page(ws, day_text):
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(size=25, bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Cabeçalho (sem borda)
    ws.merge_cells('A1:H2')
    header_cell = ws['A1']
    header_cell.value = day_text
    header_cell.alignment = align_center
    header_cell.font = bold_font

    # Largura das colunas
    for col in range(1, 9):  # Colunas A até H
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

    # Divisões de escrita (4 blocos de 9 linhas cada)
    for row in range(4, 40, 9):
        ws.merge_cells(f'A{row}:H{row+8}')
        for r in range(row, row+9):
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                cell.alignment = align_center
                ws.row_dimensions[r].height = 35


# Função principal para criar a agenda
def create_agenda():
    wb = Workbook()
    dates = generate_dates(year=ANO)

    for i, date in enumerate(dates):
        ws = wb.create_sheet(title=f"Dia {i+1}")
        setup_day_page(ws, date)

        # Configurações de impressão
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        ws.print_area = 'A1:H39'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True

    wb.remove(wb['Sheet'])
    wb.save(f'Agenda_{ANO}.xlsx')

# Executa a criação da agenda
create_agenda()
